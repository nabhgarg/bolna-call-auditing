#!/usr/bin/env python3
"""Export the local calls workbook to JSON for Supabase seeding."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
DEFAULT_INPUT = ROOT / "Copy of Calls for Auditing.xlsx"
DEFAULT_OUTPUT = ROOT / "vercel-supabase-app" / "calls.json"


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def main() -> None:
    workbook = load_workbook(DEFAULT_INPUT, read_only=True, data_only=True)
    rows = []

    for sheet in workbook.worksheets:
        iterator = sheet.iter_rows(values_only=True)
        try:
            headers = [clean(value) for value in next(iterator)]
        except StopIteration:
            continue

        header_map = {name: index for index, name in enumerate(headers)}
        if "execution_id" not in header_map:
            continue

        def get(row, column: str) -> str:
            index = header_map.get(column)
            if index is None or index >= len(row):
                return ""
            return clean(row[index])

        def get_first(row, columns: tuple[str, ...]) -> str:
            for column in columns:
                value = get(row, column)
                if value:
                    return value
            return ""

        for row in iterator:
            execution_id = get(row, "execution_id")
            if not execution_id:
                continue
            rows.append(
                {
                    "execution_id": execution_id,
                    "assigned_reviewer": get_first(row, ("assigned_reviewer", "assigned_to", "reviewer", "reviewer_name", "assignee")),
                    "org_name": get(row, "org_name"),
                    "agent_id": get(row, "agent_id"),
                    "agent_name": get(row, "agent_name"),
                    "duration_sec": float(get(row, "duration_sec") or 0),
                    "created_at_ist": get(row, "created_at_ist"),
                    "to_number": get(row, "to_number"),
                    "status": get(row, "status"),
                    "transcriber_language": get(row, "transcriber_language"),
                    "transcript": get(row, "transcript"),
                    "recording_url": get(row, "recording_url"),
                    "agent_interrupted_user_count": float(get(row, "agent_interrupted_user_count") or 0),
                    "source_sheet": sheet.title,
                }
            )

    DEFAULT_OUTPUT.write_text(json.dumps(rows, ensure_ascii=False, indent=2))
    print(f"Wrote {len(rows)} calls to {DEFAULT_OUTPUT}")


if __name__ == "__main__":
    main()
