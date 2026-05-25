#!/usr/bin/env python3
"""Small internal call-audit web tool.

Run:
    python3 audit_tool.py

Then open http://127.0.0.1:8765
"""

from __future__ import annotations

import csv
import io
import json
import mimetypes
import sqlite3
import urllib.error
import urllib.request
from datetime import datetime, timezone
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import unquote, urlparse

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
STATIC = ROOT / "static"
DB_PATH = ROOT / "audit_tool.sqlite3"
DEFAULT_CALLS_FILE = ROOT / "Copy of Calls for Auditing.xlsx"


def db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db() -> None:
    with db() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS calls (
                execution_id TEXT PRIMARY KEY,
                assigned_reviewer TEXT,
                org_name TEXT,
                agent_id TEXT,
                agent_name TEXT,
                duration_sec REAL,
                created_at_ist TEXT,
                to_number TEXT,
                status TEXT,
                transcriber_language TEXT,
                transcript TEXT,
                recording_url TEXT,
                agent_interrupted_user_count REAL,
                source_sheet TEXT,
                imported_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS reviews (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                call_id TEXT NOT NULL,
                reviewer_name TEXT,
                review_mode TEXT,
                vibe_score TEXT,
                flow_score TEXT,
                llm_rating TEXT,
                llm_error_type TEXT,
                notes TEXT,
                issues_json TEXT NOT NULL,
                started_at TEXT,
                submitted_at TEXT NOT NULL,
                duration_taken_sec INTEGER,
                sheets_synced_at TEXT,
                sheets_sync_error TEXT,
                FOREIGN KEY (call_id) REFERENCES calls(execution_id)
            );
            """
        )
        ensure_column(conn, "calls", "assigned_reviewer", "TEXT")
        ensure_column(conn, "reviews", "sheets_synced_at", "TEXT")
        ensure_column(conn, "reviews", "sheets_sync_error", "TEXT")


def ensure_column(conn: sqlite3.Connection, table: str, column: str, column_type: str) -> None:
    columns = {row["name"] for row in conn.execute(f"PRAGMA table_info({table})")}
    if column not in columns:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {column_type}")


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_env() -> dict[str, str]:
    env = {}
    path = ROOT / ".env"
    if not path.exists():
        return env
    for raw_line in path.read_text().splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip().strip('"').strip("'")
    return env


def import_calls(path: Path = DEFAULT_CALLS_FILE) -> dict:
    if not path.exists():
        raise FileNotFoundError(f"Could not find {path.name}")

    wb = load_workbook(path, read_only=True, data_only=True)
    imported = 0
    skipped = 0
    now = datetime.now(timezone.utc).isoformat()

    with db() as conn:
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            try:
                headers = [clean(v) for v in next(rows)]
            except StopIteration:
                continue

            header_map = {name: idx for idx, name in enumerate(headers)}
            required = {"execution_id", "transcript", "recording_url"}
            if not required.issubset(header_map):
                continue

            def get_first(row, cols: tuple[str, ...]) -> str:
                for col in cols:
                    idx = header_map.get(col)
                    if idx is not None and idx < len(row):
                        value = clean(row[idx])
                        if value:
                            return value
                return ""

            for row in rows:
                execution_idx = header_map["execution_id"]
                execution_id = clean(row[execution_idx] if execution_idx < len(row) else "")
                if not execution_id:
                    skipped += 1
                    continue

                def get(col: str):
                    idx = header_map.get(col)
                    if idx is None or idx >= len(row):
                        return ""
                    return clean(row[idx])

                conn.execute(
                    """
                    INSERT INTO calls (
                        execution_id, assigned_reviewer, org_name, agent_id, agent_name, duration_sec,
                        created_at_ist, to_number, status, transcriber_language,
                        transcript, recording_url, agent_interrupted_user_count,
                        source_sheet, imported_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(execution_id) DO UPDATE SET
                        assigned_reviewer=excluded.assigned_reviewer,
                        org_name=excluded.org_name,
                        agent_id=excluded.agent_id,
                        agent_name=excluded.agent_name,
                        duration_sec=excluded.duration_sec,
                        created_at_ist=excluded.created_at_ist,
                        to_number=excluded.to_number,
                        status=excluded.status,
                        transcriber_language=excluded.transcriber_language,
                        transcript=excluded.transcript,
                        recording_url=excluded.recording_url,
                        agent_interrupted_user_count=excluded.agent_interrupted_user_count,
                        source_sheet=excluded.source_sheet,
                        imported_at=excluded.imported_at
                    """,
                    (
                        execution_id,
                        get_first(row, ("assigned_reviewer", "assigned_to", "reviewer", "reviewer_name", "assignee")),
                        get("org_name"),
                        get("agent_id"),
                        get("agent_name"),
                        float(get("duration_sec") or 0),
                        get("created_at_ist"),
                        get("to_number"),
                        get("status"),
                        get("transcriber_language"),
                        get("transcript"),
                        get("recording_url"),
                        float(get("agent_interrupted_user_count") or 0),
                        ws.title,
                        now,
                    ),
                )
                imported += 1

    return {"imported": imported, "skipped": skipped, "file": path.name}


def parse_turns(transcript: str) -> list[dict]:
    turns: list[dict] = []
    current_role = None
    current_text: list[str] = []

    for raw_line in transcript.replace("\r", "\n").split("\n"):
        line = raw_line.strip()
        if not line:
            continue

        lower = line.lower()
        if lower.startswith("assistant:"):
            if current_role:
                turns.append({"role": current_role, "text": " ".join(current_text).strip()})
            current_role = "assistant"
            current_text = [line.split(":", 1)[1].strip()]
        elif lower.startswith("user:"):
            if current_role:
                turns.append({"role": current_role, "text": " ".join(current_text).strip()})
            current_role = "user"
            current_text = [line.split(":", 1)[1].strip()]
        else:
            current_text.append(line)

    if current_role:
        turns.append({"role": current_role, "text": " ".join(current_text).strip()})
    return turns


def call_summary(row: sqlite3.Row) -> dict:
    return {
        "execution_id": row["execution_id"],
        "assigned_reviewer": row["assigned_reviewer"],
        "org_name": row["org_name"],
        "agent_name": row["agent_name"],
        "duration_sec": row["duration_sec"],
        "created_at_ist": row["created_at_ist"],
        "status": row["status"],
        "language": row["transcriber_language"],
        "source_sheet": row["source_sheet"],
        "reviewed": bool(row["review_id"]),
        "reviewer_name": row["reviewer_name"],
    }


def row_to_call(row: sqlite3.Row) -> dict:
    data = dict(row)
    data["turns"] = parse_turns(data.get("transcript") or "")
    return data


def timestamp_to_seconds(timestamp: str) -> int | None:
    parts = [part.strip() for part in clean(timestamp).split(":")]
    if not parts or any(not part.isdigit() for part in parts):
        return None
    values = [int(part) for part in parts]
    if len(values) == 2:
        return values[0] * 60 + values[1]
    if len(values) == 3:
        return values[0] * 3600 + values[1] * 60 + values[2]
    return None


def recording_link_at(recording_url: str, timestamp: str) -> str:
    seconds = timestamp_to_seconds(timestamp)
    if not recording_url or seconds is None:
        return ""
    return f"{recording_url}#t={seconds}"


REVIEW_EXPORT_COLUMNS = [
    "review_id",
    "call_id",
    "org_name",
    "agent_name",
    "call_duration_sec",
    "call_created_at_ist",
    "reviewer_name",
    "review_mode",
    "vibe_score",
    "flow_score",
    "llm_rating",
    "llm_error_type",
    "notes",
    "issue_type",
    "issue_timestamp",
    "issue_recording_link",
    "issue_payload_json",
    "started_at",
    "submitted_at",
    "duration_taken_sec",
]


def review_export_rows(review_ids: list[int] | None = None, only_unsynced: bool = False) -> list[dict]:
    where = []
    params: list = []
    if review_ids:
        placeholders = ",".join("?" for _ in review_ids)
        where.append(f"r.id IN ({placeholders})")
        params.extend(review_ids)
    if only_unsynced:
        where.append("r.sheets_synced_at IS NULL")
    where_sql = f"WHERE {' AND '.join(where)}" if where else ""

    with db() as conn:
        rows = conn.execute(
            f"""
            SELECT r.*, c.org_name, c.agent_name, c.duration_sec, c.created_at_ist, c.recording_url
            FROM reviews r
            JOIN calls c ON c.execution_id = r.call_id
            {where_sql}
            ORDER BY r.call_id ASC, r.submitted_at DESC, r.id ASC
            """,
            params,
        ).fetchall()

    export_rows = []
    for row in rows:
        issues = json.loads(row["issues_json"] or "[]")
        if not issues:
            issues = [{}]
        for issue in issues:
            export_rows.append(
                {
                    "review_id": row["id"],
                    "call_id": row["call_id"],
                    "org_name": row["org_name"],
                    "agent_name": row["agent_name"],
                    "call_duration_sec": row["duration_sec"],
                    "call_created_at_ist": row["created_at_ist"],
                    "reviewer_name": row["reviewer_name"],
                    "review_mode": row["review_mode"],
                    "vibe_score": row["vibe_score"],
                    "flow_score": row["flow_score"],
                    "llm_rating": row["llm_rating"],
                    "llm_error_type": row["llm_error_type"],
                    "notes": row["notes"],
                    "issue_type": issue.get("type", ""),
                    "issue_timestamp": issue.get("timestamp", ""),
                    "issue_recording_link": recording_link_at(row["recording_url"], issue.get("timestamp", "")),
                    "issue_payload_json": json.dumps(issue, ensure_ascii=False),
                    "started_at": row["started_at"],
                    "submitted_at": row["submitted_at"],
                    "duration_taken_sec": row["duration_taken_sec"],
                }
            )
    return export_rows


def sync_reviews_to_sheets(review_ids: list[int] | None = None) -> dict:
    webhook_url = load_env().get("GOOGLE_SHEETS_WEBHOOK_URL", "")
    if not webhook_url:
        return {"ok": False, "configured": False, "synced_reviews": 0, "error": "GOOGLE_SHEETS_WEBHOOK_URL is not set in .env"}

    rows = review_export_rows(review_ids=review_ids, only_unsynced=review_ids is None)
    if not rows:
        return {"ok": True, "configured": True, "synced_reviews": 0, "rows": 0}

    payload = json.dumps({"columns": REVIEW_EXPORT_COLUMNS, "rows": rows}, ensure_ascii=False).encode("utf-8")
    request = urllib.request.Request(
        webhook_url,
        data=payload,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    review_id_values = sorted({int(row["review_id"]) for row in rows})
    now = datetime.now(timezone.utc).isoformat()
    try:
        with urllib.request.urlopen(request, timeout=15) as response:
            response_body = response.read().decode("utf-8")
            if response.status >= 400:
                raise RuntimeError(response_body or f"HTTP {response.status}")
        with db() as conn:
            conn.executemany(
                "UPDATE reviews SET sheets_synced_at = ?, sheets_sync_error = NULL WHERE id = ?",
                [(now, review_id) for review_id in review_id_values],
            )
        return {"ok": True, "configured": True, "synced_reviews": len(review_id_values), "rows": len(rows)}
    except (urllib.error.URLError, TimeoutError, RuntimeError) as exc:
        error = str(exc)
        with db() as conn:
            conn.executemany(
                "UPDATE reviews SET sheets_sync_error = ? WHERE id = ?",
                [(error, review_id) for review_id in review_id_values],
            )
        return {"ok": False, "configured": True, "synced_reviews": 0, "rows": len(rows), "error": error}


class Handler(BaseHTTPRequestHandler):
    server_version = "BolnaAuditTool/0.1"

    def log_message(self, fmt, *args):
        return

    def send_json(self, payload, status=HTTPStatus.OK):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def read_json(self) -> dict:
        length = int(self.headers.get("Content-Length", "0"))
        if length == 0:
            return {}
        return json.loads(self.rfile.read(length).decode("utf-8"))

    def do_GET(self):
        parsed = urlparse(self.path)
        path = unquote(parsed.path)

        if path == "/":
            return self.serve_file(STATIC / "index.html")
        if path.startswith("/static/"):
            return self.serve_file(ROOT / path.lstrip("/"))
        if path == "/api/calls":
            return self.api_calls()
        if path.startswith("/api/calls/"):
            return self.api_call(path.rsplit("/", 1)[-1])
        if path == "/api/reviews.csv":
            return self.api_reviews_csv()
        self.send_error(HTTPStatus.NOT_FOUND)

    def do_POST(self):
        parsed = urlparse(self.path)
        path = unquote(parsed.path)
        if path == "/api/import":
            try:
                result = import_calls()
                return self.send_json(result)
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
        if path == "/api/reviews":
            return self.api_save_review()
        if path == "/api/sync-sheets":
            return self.api_sync_sheets()
        self.send_error(HTTPStatus.NOT_FOUND)

    def serve_file(self, path: Path):
        if not path.exists() or not path.is_file():
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        body = path.read_bytes()
        content_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def api_calls(self):
        with db() as conn:
            rows = conn.execute(
                """
                SELECT c.*, r.id AS review_id, r.reviewer_name
                FROM calls c
                LEFT JOIN reviews r ON r.call_id = c.execution_id
                ORDER BY COALESCE(c.created_at_ist, '') DESC, c.execution_id
                """
            ).fetchall()
        return self.send_json({"calls": [call_summary(row) for row in rows]})

    def api_call(self, execution_id: str):
        with db() as conn:
            row = conn.execute(
                "SELECT * FROM calls WHERE execution_id = ?", (execution_id,)
            ).fetchone()
            if not row:
                return self.send_json({"error": "Call not found"}, HTTPStatus.NOT_FOUND)
            reviews = conn.execute(
                "SELECT * FROM reviews WHERE call_id = ? ORDER BY submitted_at DESC",
                (execution_id,),
            ).fetchall()
        payload = row_to_call(row)
        payload["reviews"] = [dict(r) for r in reviews]
        return self.send_json(payload)

    def api_save_review(self):
        payload = self.read_json()
        call_id = clean(payload.get("call_id"))
        if not call_id:
            return self.send_json({"error": "call_id is required"}, HTTPStatus.BAD_REQUEST)

        issues = payload.get("issues") or []
        now = datetime.now(timezone.utc).isoformat()
        with db() as conn:
            exists = conn.execute(
                "SELECT 1 FROM calls WHERE execution_id = ?", (call_id,)
            ).fetchone()
            if not exists:
                return self.send_json({"error": "Call not found"}, HTTPStatus.NOT_FOUND)
            cursor = conn.execute(
                """
                INSERT INTO reviews (
                    call_id, reviewer_name, review_mode, vibe_score, flow_score,
                    llm_rating, llm_error_type, notes, issues_json, started_at,
                    submitted_at, duration_taken_sec
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    call_id,
                    clean(payload.get("reviewer_name")),
                    clean(payload.get("review_mode")),
                    clean(payload.get("vibe_score")),
                    clean(payload.get("flow_score")),
                    clean(payload.get("llm_rating")),
                    clean(payload.get("llm_error_type")),
                    clean(payload.get("notes")),
                    json.dumps(issues, ensure_ascii=False),
                    clean(payload.get("started_at")),
                    now,
                    int(payload.get("duration_taken_sec") or 0),
                ),
            )
        sync_result = sync_reviews_to_sheets([cursor.lastrowid])
        return self.send_json({"ok": True, "review_id": cursor.lastrowid, "sheets_sync": sync_result})

    def api_reviews_csv(self):
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(REVIEW_EXPORT_COLUMNS)
        for row in review_export_rows():
            writer.writerow([row.get(column, "") for column in REVIEW_EXPORT_COLUMNS])

        body = output.getvalue().encode("utf-8-sig")
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "text/csv; charset=utf-8")
        self.send_header("Content-Disposition", 'attachment; filename="bolna_call_reviews.csv"')
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def api_sync_sheets(self):
        result = sync_reviews_to_sheets()
        return self.send_json(result, HTTPStatus.OK if result.get("ok") else HTTPStatus.BAD_REQUEST)


def ensure_imported() -> None:
    with db() as conn:
        count = conn.execute("SELECT COUNT(*) FROM calls").fetchone()[0]
    if count == 0 and DEFAULT_CALLS_FILE.exists():
        import_calls(DEFAULT_CALLS_FILE)


def main() -> None:
    init_db()
    ensure_imported()
    port = 8765
    server = ThreadingHTTPServer(("127.0.0.1", port), Handler)
    print(f"Call audit tool running at http://127.0.0.1:{port}")
    print("Press Ctrl+C to stop.")
    server.serve_forever()


if __name__ == "__main__":
    main()
